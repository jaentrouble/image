import openpyxl as xl
import joblib
from sklearn.ensemble import RandomForestClassifier
from common.constants import *
import os

wb = xl.load_workbook(filename = AUTO_database_filename)
ws = wb.worksheets[0]
row = 1
x = []
y = []
while ws.cell(row, 1).value != None :
    v = []
    for i in range(AUTO_vector_size) :
        v.append(ws.cell(row, i+1).value)
    x.append(v)
    y.append(ws.cell(row,AUTO_vector_size+1).value)
    row += 1

clf = RandomForestClassifier()
clf = clf.fit(x,y)
joblib.dump(clf, os.path.join(AUTO_PATH, AUTO_RFC_filename))
print('model saved')