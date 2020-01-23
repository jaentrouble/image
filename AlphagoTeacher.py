import openpyxl as xl
import tensorflow as tf
import numpy as np
from common.constants import *
import os

wb = xl.load_workbook(filename = AUTO_database_filename)
ws = wb.worksheets[0]
row = 1
x = []
y = []
while ws.cell(row, 1).value != None :
    v = []
    for i in range(AUTO_vector_size) :
        v.append(ws.cell(row, i+1).value)
    x.append(v)
    y.append(ws.cell(row,AUTO_vector_size+1).value)
    row += 1

model = tf.keras.models.load_model(os.path.join(AUTO_PATH, AUTO_default_filename))
model.fit(x = x, y = y, epochs = AUTO_teacher_epoch)
model.save(os.path.join(AUTO_PATH, AUTO_default_filename))