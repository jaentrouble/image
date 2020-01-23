import openpyxl as xl

class Saver() :
    def __init__(self, filename = 'savefile.xlsx') :
        self.filename = filename
        try :
            self.wb = xl.load_workbook(filename = self.filename)
        except FileNotFoundError :
            self.wb = xl.Workbook()
            self.wb.save(self.filename)
        self.ws = self.wb.worksheets[0]
        self.rec_col = 1
        self.row = 1

    def save(self, record_list : list):
        """
        record list : [[a,b, ...], ...] saves a, b, ... to each column
        """
        try :
            while self.ws.cell(self.row,self.rec_col).value != None :
                self.row += 1
            for idx in range(len(record_list)) :
                for col in range(len(record_list[idx])) :
                    self.ws.cell(self.row, self.rec_col + col).value = record_list[idx][col]
                self.row += 1
            self.wb.save(self.filename)
        except :
            return False
        return True